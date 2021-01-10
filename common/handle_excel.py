"""
Name:   68
Email:  liubo3398@163.com
Time:   2020/12/6 10:34
"""
import openpyxl


class HandleExcel:
    """操作excel封装"""

    def __init__(self, workname, sheetname):
        """
        :param workname: Excel文件名
        :param sheetname: 表单名
        """
        self.workname = workname
        self.sheetname = sheetname

    # 读取文件
    def read_excel(self):
        """
        :return: cases:用例数据 列表套字典的形式
        """
        work = openpyxl.load_workbook(self.workname)
        sh1 = work[self.sheetname]
        rows = list(sh1.rows)

        title = [i.value for i in rows[0]]
        cases = []
        for i in rows[1:]:
            row = []
            for j in i:
                row.append(j.value)
            cases.append(dict(zip(title, row)))
        return cases

    # 写入文件
    def write_excel(self, row, column, value):
        """
        :param row:     int 行号
        :param column:  int 列号
        :param value:   str 写入的内容
        :return: None
        """
        work = openpyxl.load_workbook(self.workname)
        sh1 = work[self.sheetname]
        sh1.cell(row=row, column=column, value=value)
        work.save(self.workname)


if __name__ == '__main__':
    # a = HandleExcel('../datas/data.xlsx', 'register')
    # cases = a.read_excel()
    # print(cases)
    pass